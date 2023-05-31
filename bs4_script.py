import openpyxl
import requests
from bs4 import BeautifulSoup
import openpyxl
from concurrent.futures import ThreadPoolExecutor

class WebScraper:
    def scrape_university_details(self, link):
        try:
            response = requests.get(link)
            soup = BeautifulSoup(response.content, 'html.parser')

            whed_link = link
            uni_name = None
            street = None
            city = None
            province = None
            post_code = None
            website = None
            statistics_year = None
            total_student = None
            total_staff = None
            institution_funding = None
            history = None
            academic_year = None
            language = None
            bachelor_degree = None
            master_degree = None
            doctor_degree = None
            diploma_certificate = None

            uni_country_element = soup.select_one('.country')
            uni_country = uni_country_element.text if uni_country_element else None

            uni_name_element = soup.select_one('#page > div > div > div.detail_right > div')
            uni_name = uni_name_element.text.strip() if uni_name_element else None

            libelle_elements = soup.select('.libelle')
            for item in libelle_elements:
                try:
                    if item.text == 'Street:':
                        street_element = item.find_next_sibling('span')
                        street = street_element.text.strip() if street_element else None
                    elif item.text == 'City:':
                        city_element = item.find_next_sibling('span')
                        city = city_element.text.strip() if city_element else None
                    elif item.text == 'Province:':
                        province_element = item.find_next_sibling('span')
                        province = province_element.text.strip() if province_element else None
                    elif item.text == 'Post Code:':
                        post_code_element = item.find_next_sibling('span')
                        post_code = post_code_element.text.strip() if post_code_element else None
                    elif item.text == 'WWW:':
                        website_element = item.find_next_sibling('span')
                        website = website_element.text.strip() if website_element else None
                except Exception as e:
                    print(f"Error scraping {item.text}: {str(e)}")

            dt_elements = soup.select('#contenu > .dl > .dt')
            for item in dt_elements:
                try:
                    if item.text == 'Institution Funding':
                        institution_funding_element = item.find_next('div')
                        institution_funding = institution_funding_element.text.strip() if institution_funding_element else None
                    elif item.text == 'History':
                        history_element = item.find_next('div')
                        history = history_element.text.strip() if history_element else None
                    elif item.text == 'Academic Year':
                        academic_year_element = item.find_next('div')
                        academic_year = academic_year_element.text.strip() if academic_year_element else None
                    elif item.text == 'Language(s)':
                        language_element = item.find_next('div')
                        language = language_element.text.strip() if language_element else None
                    elif item.text == 'Staff':
                        statistics_year_element = soup.find('span', text='Statistics Year:').find_next_sibling('span')
                        statistics_year = statistics_year_element.text.strip()
                        total_staff_element = soup.find('span', text='Full Time Total:').find_next_sibling('span')
                        total_staff = total_staff_element.text.strip()
                    elif item.text == 'Students':
                        total_student_element = soup.find('span', text='Total:').find_next_sibling('span')
                        total_student = total_student_element.text.strip()
                except Exception as e:
                    print(f"Error scraping {item.text}: {str(e)}")

            principal_elements = soup.select('.principal')
            for item in principal_elements:
                try:
                    if item.text == "Bachelor's Degree":
                        bachelor_element = item.find_next_sibling('p').find('span', class_='contenu')
                        bachelor_degree = bachelor_element.text.strip() if bachelor_element else None
                    elif item.text == "Master's Degree":
                        master_element = item.find_next_sibling('p').find('span', class_='contenu')
                        master_degree = master_element.text.strip() if master_element else None
                    elif item.text == "Doctor's Degree (Research/Scholarship)":
                        doctor_element = item.find_next_sibling('p').find('span', class_='contenu')
                        doctor_degree = doctor_element.text.strip() if doctor_element else None
                    elif item.text == "Post-bachelor's Diploma/Certificate":
                        diploma_element = item.find_next_sibling('p').find('span', class_='contenu')
                        diploma_certificate = diploma_element.text.strip() if diploma_element else None
                except Exception as e:
                    print(f"Error scraping {item.text}: {str(e)}")

            try:
                workbook = openpyxl.load_workbook('university_data.xlsx')
            except FileNotFoundError:
                workbook = openpyxl.Workbook()

            sheet = workbook.active
            last_row = sheet.max_row

            if last_row == 1:
                sheet.append(('Whed Link', 'University Name', 'Country', 'Street', 'City', 'Province', 'Post Code', 'Website',
                            'Statistics Year', 'Total Staff', 'Total Student', 'Institution Funding', 'History', 'Academic Year',
                            'Language(s)', "Bachelor's Degree", "Master's Degree", "Doctor's Degree", 'Diploma/Certificate'))

            sheet.append((whed_link, uni_name, uni_country, street, city, province, post_code, website, statistics_year, total_staff, 
                        total_student, institution_funding, history, academic_year, language, bachelor_degree, master_degree, doctor_degree, diploma_certificate))

            workbook.save('university_data.xlsx')

        except Exception as e:
            try:
                workbook2 = openpyxl.load_workbook('declined_links.xlsx')
            except FileNotFoundError:
                workbook2 = openpyxl.Workbook()
            sheet = workbook2.active
            sheet.append([link])
            workbook2.save('declined_links.xlsx')

    def run_scraper(self):
        try:
            workbook = openpyxl.load_workbook('whed_uni_links.xlsx')
            sheet = workbook.active
            links_column = sheet['A']

            total_links = len(links_column)
            i = 10192
            with ThreadPoolExecutor(max_workers=5) as executor:
                for cell in links_column[10192:]:
                    link = cell.value
                    executor.submit(self.scrape_university_details(link))
                    i += 1
                    print(f"Progress: {round((i*100)/total_links, 2)}%  {i}/{total_links}")

        except Exception as e:
            print("Error:", str(e))


scraper = WebScraper()
scraper.run_scraper()
