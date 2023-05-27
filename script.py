import math
import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import Select
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import chromedriver_autoinstaller
import openpyxl

class WebScraper:
    def __init__(self):
        self.driver = None
        self.wait = None

    def setup_driver(self):
        chromedriver_autoinstaller.install()

        options = Options()
        options.add_argument("--headless")

        self.driver = webdriver.Chrome(options=options)
        self.wait = WebDriverWait(self.driver, 10)

    def close_driver(self):
        if self.driver:
            self.driver.quit()

    def accept_cookie(self):
        try:
            accept_cookie = self.wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, '#accord_cookie > center > input.bouton')))
            accept_cookie.click()
        except Exception as e:
            print("Error accepting cookie:", str(e))

    def select_country(self, index):
        try:
            country_select = Select(self.wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, '#Chp1'))))
            options = country_select.options
            button = self.wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, '#fsearch > p > input[type=button]')))
            country_name = options[index].text

            country_select.select_by_index(index)
            button.click()

            return country_name
        except Exception as e:
            print("Error selecting country:", str(e))
            return None

    def scrape_unique_links(self, country_name):
        try:
            info_element = self.wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, 'p.infos')))
            result_count = info_element.text.split()[0]
            try:
                result_count = int(result_count)
                print(f"Found {country_name} - Total count: {result_count}")
            except ValueError:
                print(f"Skipping {country_name} - result count: 0")
                return
            
            if result_count % 10 != 0:
                result_count = math.ceil(result_count / 10)

            for i in range(result_count):
                progress = f"Scraping {country_name}: {i + 1}/{result_count}"
                print(progress, end="\r")

                uni_cards = self.wait.until(EC.presence_of_all_elements_located((By.CSS_SELECTOR, 'div.details > h3 > a')))
                for uni_card in uni_cards:
                    try:
                        link = uni_card.get_attribute('href')
                        self.scrape_university_details(link)
                    except Exception as e:
                        print("Error scraping university details:", str(e))
                        
                self.driver.switch_to.window(self.driver.window_handles[0])
                self.navigate_to_next_page()

        except Exception as e:
            print("Error scraping university data:", str(e))

    def scrape_university_details(self, link):
        try:
            self.driver.execute_script("window.open();")
            self.driver.switch_to.window(self.driver.window_handles[1])
            self.driver.get(link)

            whed_link = None
            uni_name = None
            street = None
            city = None
            province = None
            post_code = None
            website = None
            statistics_year = None
            total_student = None
            total_staff = None
            institution_funding = None
            history = None
            academic_year = None
            language = None
            bachelor_degree = None
            master_degree = None
            doctor_degree = None
            diploma_certificate = None

            uni_country_element = self.wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, '.country')))
            uni_country = uni_country_element.text

            whed_link = link

            uni_name_element = self.wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, '.detail_right > div')))
            uni_name = uni_name_element.text

            libelle_elements = self.wait.until(EC.presence_of_all_elements_located((By.CSS_SELECTOR, '.libelle')))
            for item in libelle_elements:
                if item.text == 'Street:':
                    parent_node = item.find_element(By.XPATH, '..')
                    child_elements = parent_node.find_elements(By.XPATH, '*')
                    street_element = child_elements[child_elements.index(item)+1]
                    street = street_element.text
                elif item.text == 'City:':
                    parent_node = item.find_element(By.XPATH, '..')
                    child_elements = parent_node.find_elements(By.XPATH, '*')
                    city_element = child_elements[child_elements.index(item)+1]
                    city = city_element.text
                elif item.text == 'Province:':
                    parent_node = item.find_element(By.XPATH, '..')
                    child_elements = parent_node.find_elements(By.XPATH, '*')
                    province_element = child_elements[child_elements.index(item)+1]
                    province = province_element.text
                elif item.text == 'Post Code:':
                    parent_node = item.find_element(By.XPATH, '..')
                    child_elements = parent_node.find_elements(By.XPATH, '*')
                    post_code_element = child_elements[child_elements.index(item)+1]
                    post_code = post_code_element.text
                elif item.text == 'WWW:':
                    parent_node = item.find_element(By.XPATH, '..')
                    child_elements = parent_node.find_elements(By.XPATH, '*')
                    website_element = child_elements[child_elements.index(item)+1]
                    website = website_element.text

            dl_elements = self.wait.until(EC.presence_of_all_elements_located((By.CSS_SELECTOR, '#contenu > .dl > .dt')))
            for item in dl_elements:
                if item.text == 'Institution Funding':
                    parent_node = item.find_element(By.XPATH, '..')
                    second_child = parent_node.find_elements(By.XPATH, '*')[1]
                    institution_funding_element = second_child.find_elements(By.XPATH, '*')[0]
                    institution_funding = institution_funding_element.text
                elif item.text == 'History':
                    parent_node = item.find_element(By.XPATH, '..')
                    second_child = parent_node.find_elements(By.XPATH, '*')[1]
                    history_element = second_child.find_elements(By.XPATH, '*')[0]
                    history = history_element.text
                elif item.text == 'Academic Year':
                    parent_node = item.find_element(By.XPATH, '..')
                    second_child = parent_node.find_elements(By.XPATH, '*')[1]
                    academic_year_element = second_child.find_elements(By.XPATH, '*')[0]
                    academic_year = academic_year_element.text
                elif item.text == 'Language(s)':
                    parent_node = item.find_element(By.XPATH, '..')
                    second_child = parent_node.find_elements(By.XPATH, '*')[1]
                    language_element = second_child.find_elements(By.XPATH, '*')[0]
                    language = language_element.text
                elif item.text == 'History':
                    parent_node = item.find_element(By.XPATH, '..')
                    second_child = parent_node.find_elements(By.XPATH, '*')[1]
                    history_element = second_child.find_elements(By.XPATH, '*')[0]
                    history = history_element.text
                elif item.text == 'Staff':
                    parent_node = item.find_element(By.XPATH, '..')
                    second_child = parent_node.find_elements(By.XPATH, '*')[1]
                    second_child_list = second_child.find_elements(By.XPATH, '*')
                    for item in second_child_list:
                        item_list =  item.find_elements(By.XPATH, '*')
                        try: 
                            if item_list[0].text == 'Statistics Year:':
                                statistics_year = item_list[1].text
                            elif item_list[0].text == 'Full Time Total:':
                                total_staff = item_list[1].text
                        except:
                            pass
                elif item.text == 'Students':
                    parent_node = item.find_element(By.XPATH, '..')
                    second_child = parent_node.find_elements(By.XPATH, '*')[1]
                    second_child_list = second_child.find_elements(By.XPATH, '*')
                    for item in second_child_list:
                        item_list =  item.find_elements(By.XPATH, '*')
                        try: 
                            if item_list[0].text == 'Total:':
                                total_student = item_list[1].text
                        except:
                            pass

            principal_elements = self.wait.until(EC.presence_of_all_elements_located((By.CSS_SELECTOR, '.principal')))
            for item in principal_elements:
                if item.text == "Bachelor's Degree":
                    parent_node = item.find_element(By.XPATH, '..')
                    second_child = parent_node.find_elements(By.XPATH, '*')[1]
                    bachelor_element = second_child.find_elements(By.XPATH, '*')[1]
                    bachelor_degree = bachelor_element.text
                elif item.text == "Master's Degree":
                    parent_node = item.find_element(By.XPATH, '..')
                    second_child = parent_node.find_elements(By.XPATH, '*')[1]
                    master_element = second_child.find_elements(By.XPATH, '*')[1]
                    master_degree = master_element.text
                elif item.text == "Doctor's Degree (Research/Scholarship)":
                    parent_node = item.find_element(By.XPATH, '..')
                    second_child = parent_node.find_elements(By.XPATH, '*')[1]
                    doctor_element = second_child.find_elements(By.XPATH, '*')[1]
                    doctor_degree = doctor_element.text
                elif item.text == "Post-bachelor's Diploma/Certificate":
                    parent_node = item.find_element(By.XPATH, '..')
                    second_child = parent_node.find_elements(By.XPATH, '*')[1]
                    diploma_element = second_child.find_elements(By.XPATH, '*')[1]
                    diploma_certificate = diploma_element.text

            try:
                workbook = openpyxl.load_workbook('university_data.xlsx')
            except FileNotFoundError:
                workbook = openpyxl.Workbook()

            sheet = workbook.active
            last_row = sheet.max_row

            if last_row == 1:
                sheet.append(('Whed Link', 'University Name', 'Country', 'Street', 'City', 'Province', 'Post Code', 'Website',
                            'Statistics Year', 'Total Staff', 'Total Student', 'Institution Funding', 'History', 'Academic Year',
                            'Language(s)', "Bachelor's Degree", "Master's Degree", "Doctor's Degree", 'Diploma/Certificate'))

            sheet.append((whed_link, uni_name, uni_country, street, city, province, post_code, website, statistics_year, total_staff, 
                        total_student, institution_funding, history, academic_year, language, bachelor_degree, master_degree, doctor_degree, diploma_certificate))

            workbook.save('university_data.xlsx')
            
            self.driver.close()

            self.driver.switch_to.window(self.driver.window_handles[0])
        except Exception as e:
            print("Error scraping university details:", str(e))

    def navigate_to_next_page(self):
        try:
            next_button = self.wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, '.next')))
            next_button.click()
        except Exception as e:
            print("Error navigating to next page:", str(e))

    def run_scraper(self):
        try:
            self.setup_driver()

            self.driver.get("https://www.whed.net/")
            self.accept_cookie()

            skip_indices = [0, 18, 19, 20, 36, 37, 38, 39, 40, 41, 42, 43, 44, 45, 46, 47, 48, 207, 208, 209, 210, 211, 212, 213, 214, 215, 216, 217, 218, 219, 220, 221, 222, 223, 224, 225, 226, 227, 228, 229, 230, 231, 232, 233, 234, 235, 236, 237, 238, 239, 240, 241, 242, 243, 244, 245, 246, 247, 248, 249, 250, 251, 252, 253, 254, 255, 256, 257, 258, 259, 260, 261]
            country_select = Select(self.wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, '#Chp1'))))
            options = country_select.options

            for i in range(len(options)):
                if i in skip_indices:
                    continue

                country_name = self.select_country(i)
                if country_name:
                    self.scrape_unique_links(country_name)

        except Exception as e:
            print("Error running the scraper:", str(e))

        finally:
            self.close_driver()

scraper = WebScraper()
scraper.run_scraper()
